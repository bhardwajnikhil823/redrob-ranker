#!/usr/bin/env python3
"""
export_submission.py - turn the ranked submission.csv into a recruiter-facing
XLSX shortlist (required by the submission portal).

It joins the top-100 ranking with each candidate's profile (name, title, company,
location, experience) from the candidate pool, so the spreadsheet reads as an
actual shortlist rather than a list of opaque IDs. Falls back gracefully to the
raw CSV columns if the pool file is not available.

Usage:
    python export_submission.py \
        --submission submission.csv \
        --candidates candidates.jsonl \
        --out submission.xlsx
"""

from __future__ import annotations

import argparse
import csv
import gzip
import io
import json
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Columns in the output sheet (header -> width).
COLUMNS = [
    ("rank", 6),
    ("candidate_id", 15),
    ("name", 20),
    ("current_title", 28),
    ("current_company", 22),
    ("location", 20),
    ("country", 12),
    ("years_experience", 12),
    ("score", 10),
    ("reasoning", 90),
]


def _open_any(path: str):
    with open(path, "rb") as probe:
        magic = probe.read(2)
    if path.endswith(".gz") or magic == b"\x1f\x8b":
        return io.TextIOWrapper(gzip.open(path, "rb"), encoding="utf-8")
    return open(path, "r", encoding="utf-8")


def load_profiles(candidates_path: str, wanted: set) -> Dict[str, Dict[str, Any]]:
    """Return {candidate_id: profile-ish dict} for just the wanted IDs."""
    out: Dict[str, Dict[str, Any]] = {}
    try:
        with _open_any(candidates_path) as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                # Cheap pre-filter avoids json.loads on every one of 100K lines.
                cid_guess = line[:40]
                if not any(w in cid_guess for w in wanted):
                    if not any(w in line for w in wanted):
                        continue
                c = json.loads(line)
                cid = c.get("candidate_id")
                if cid in wanted:
                    p = c.get("profile", {}) or {}
                    out[cid] = {
                        "name": p.get("anonymized_name", ""),
                        "current_title": p.get("current_title", ""),
                        "current_company": p.get("current_company", ""),
                        "location": p.get("location", ""),
                        "country": p.get("country", ""),
                        "years_experience": p.get("years_of_experience", ""),
                    }
                    if len(out) == len(wanted):
                        break
    except OSError:
        pass  # pool not present - we'll just emit the CSV columns
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--submission", default="submission.csv")
    ap.add_argument("--candidates", default="candidates.jsonl")
    ap.add_argument("--out", default="submission.xlsx")
    args = ap.parse_args()

    # Read the ranked CSV.
    with open(args.submission, "r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    wanted = {r["candidate_id"] for r in rows}
    profiles = load_profiles(args.candidates, wanted)

    wb = Workbook()
    ws = wb.active
    ws.title = "Top 100 Candidates"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(vertical="top", wrap_text=True)

    # Header row.
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows.
    for r_idx, row in enumerate(rows, start=2):
        cid = row["candidate_id"]
        prof = profiles.get(cid, {})
        values = {
            "rank": int(row["rank"]),
            "candidate_id": cid,
            "name": prof.get("name", ""),
            "current_title": prof.get("current_title", ""),
            "current_company": prof.get("current_company", ""),
            "location": prof.get("location", ""),
            "country": prof.get("country", ""),
            "years_experience": prof.get("years_experience", ""),
            "score": float(row["score"]),
            "reasoning": row.get("reasoning", ""),
        }
        for col_idx, (name, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=values[name])
            cell.alignment = wrap
            if name == "score":
                cell.number_format = "0.000000"

    # Freeze header + first two columns, add autofilter.
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

    wb.save(args.out)
    print(f"Wrote {len(rows)} ranked candidates to {args.out} "
          f"({len(profiles)} enriched with profile data).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
